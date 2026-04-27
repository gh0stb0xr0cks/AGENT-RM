"""
00_extract_mitre_xlsx.py — Extraction des feuilles MITRE ATT&CK depuis xlsx → JSON.

Pour chaque fichier `corpus/raw/mitre/*.xlsx` (enterprise, ics, mobile),
extrait les feuilles canoniques utiles à l'enrichissement de l'atelier 4
(scénarios opérationnels) :

    tactics · techniques · software · groups · campaigns · mitigations

Sortie :
    corpus/raw/mitre/{matrix}__{sheet}.json   (un fichier par couple matrice × feuille)
    corpus/raw/mitre/mitre_index.json         (manifeste : matrices, feuilles, counts)

Ces JSON sont ensuite consommés par `01_extract_pdf.py` qui les transforme
en chunks textuels et les ajoute à `corpus/raw/index.jsonl`.

Usage :
    python 00_extract_mitre_xlsx.py
    python 00_extract_mitre_xlsx.py --only enterprise-attack-v18.1.xlsx
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("pip install openpyxl")


ROOT = Path(__file__).resolve().parents[1]
MITRE_DIR = ROOT / "raw" / "mitre"

# Feuilles à extraire — alignées sur l'usage atelier 4.
TARGET_SHEETS: list[str] = [
    "tactics",
    "techniques",
    "software",
    "groups",
    "campaigns",
    "mitigations",
]


# ---------------------------------------------------------------------------
def to_snake(name: str) -> str:
    """`'STIX ID'` → `'stix_id'`, `'is sub-technique'` → `'is_sub_technique'`."""
    s = name.strip().lower()
    s = re.sub(r"[\s\-]+", "_", s)
    s = re.sub(r"[^a-z0-9_]", "", s)
    s = re.sub(r"_+", "_", s)
    return s.strip("_")


def normalise_value(v):
    """Convertit les valeurs openpyxl en JSON-friendly."""
    if v is None:
        return None
    if isinstance(v, (str, int, float, bool)):
        if isinstance(v, str):
            stripped = v.strip()
            return stripped if stripped else None
        return v
    return str(v)


def extract_sheet(ws) -> list[dict]:
    """Extrait une feuille en liste de dicts, en-têtes normalisées en snake_case."""
    rows = ws.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        return []

    keys = [to_snake(h) if h else f"col_{i}" for i, h in enumerate(header)]

    records: list[dict] = []
    for row in rows:
        if not any(cell is not None and str(cell).strip() for cell in row):
            continue  # ligne vide
        record = {k: normalise_value(v) for k, v in zip(keys, row)}
        records.append(record)
    return records


def matrix_stem(xlsx_path: Path) -> str:
    """`enterprise-attack-v18.1.xlsx` → `enterprise-attack-v18.1`."""
    return xlsx_path.stem


def process_workbook(xlsx_path: Path) -> dict:
    """Ouvre un xlsx, extrait les feuilles cibles, écrit un JSON par feuille."""
    print(f"\n[XLSX] {xlsx_path.name}")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    stem = matrix_stem(xlsx_path)

    summary = {
        "matrix_file": xlsx_path.name,
        "matrix_stem": stem,
        "sheets": {},
    }

    try:
        for sheet_name in TARGET_SHEETS:
            if sheet_name not in wb.sheetnames:
                print(f"  [SKIP] feuille absente : {sheet_name}")
                continue

            ws = wb[sheet_name]
            records = extract_sheet(ws)
            out_path = MITRE_DIR / f"{stem}__{sheet_name}.json"
            out_path.write_text(
                json.dumps(records, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
            summary["sheets"][sheet_name] = {
                "file": out_path.name,
                "count": len(records),
            }
            print(f"  → {sheet_name:<12} {len(records):>5} entrées  ({out_path.name})")
    finally:
        wb.close()

    return summary


# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(
        description="Extraction MITRE ATT&CK xlsx → JSON pour l'atelier 4"
    )
    parser.add_argument(
        "--only",
        help="Limite l'extraction à un fichier xlsx précis (nom de fichier)",
    )
    args = parser.parse_args()

    if not MITRE_DIR.exists():
        sys.exit(f"Répertoire introuvable : {MITRE_DIR}")

    xlsx_files = sorted(
        f for f in MITRE_DIR.glob("*.xlsx") if not f.name.startswith(".~")
    )
    if args.only:
        xlsx_files = [f for f in xlsx_files if f.name == args.only]
        if not xlsx_files:
            sys.exit(f"Aucun xlsx ne correspond à --only={args.only}")

    if not xlsx_files:
        sys.exit(f"Aucun .xlsx trouvé dans {MITRE_DIR}")

    summaries = [process_workbook(f) for f in xlsx_files]

    index_path = MITRE_DIR / "mitre_index.json"
    index_path.write_text(
        json.dumps(
            {
                "schema_version": 1,
                "target_sheets": TARGET_SHEETS,
                "matrices": summaries,
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    total = sum(
        s["count"] for m in summaries for s in m["sheets"].values()
    )
    print(f"\nManifeste : {index_path}")
    print(f"Total : {total} entrées extraites sur {len(xlsx_files)} matrice(s).")


if __name__ == "__main__":
    main()
